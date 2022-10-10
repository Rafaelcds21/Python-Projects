from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Problemas: - fazer o usuário voltar passos no meio do processo;
# - melhorar a dinâmica das respostas dos usuários (por enquanto ok)
# - Não está aceitando resposta negativa na parte de pergunta sobre os gráficos (linhas: 262 e 542)

# Permite que o usuário digite os dados manualmente


def digitar_dados():
    contador = 0
    soma = 0
    lista = []
    variancia_sem_quadrado = 0
    m = 0

    def histograma():
        cor = input('Digite a cor do gráfico: ')

        cor_ingles_preto = cor.replace(f'{cor}', 'black')
        cor_ingles_branco = cor.replace(f'{cor}', 'white')
        cor_ingles_vermelho = cor.replace(f'{cor}', 'red')
        cor_ingles_azul = cor.replace(f'{cor}', 'blue')
        cor_ingles_verde = cor.replace(f'{cor}', 'green')
        cor_ingles_amarelo = cor.replace(f'{cor}', 'yellow')

        cor_da_borda = input('Digite a cor da borda: ')

        cor_da_borda_ingles_preto = cor.replace(f'{cor}', 'black')
        cor_da_borda_ingles_branco = cor.replace(f'{cor}', 'white')
        cor_da_borda_ingles_vermelho = cor.replace(f'{cor}', 'red')
        cor_da_borda_ingles_azul = cor.replace(f'{cor}', 'blue')
        cor_da_borda_ingles_verde = cor.replace(f'{cor}', 'green')
        cor_da_borda_ingles_amarelo = cor.replace(f'{cor}', 'yellow')

        titulo = input('Digite o título do gráfico: ')
        titulo_eixox = input('Digite o título do eixo x: ')
        titulo_eixoy = input('Digite o título do eixo y: ')

        if cor == 'preto' and cor_da_borda == 'preto':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'preto' and cor_da_borda == 'branco':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'preto' and cor_da_borda == 'vermelho':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'preto' and cor_da_borda == 'azul':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'preto' and cor_da_borda == 'verde':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'preto' and cor_da_borda == 'amarelo':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        if cor == 'branco' and cor_da_borda == 'preto':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'branco' and cor_da_borda == 'branco':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'branco' and cor_da_borda == 'vermelho':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'branco' and cor_da_borda == 'azul':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'branco' and cor_da_borda == 'verde':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'branco' and cor_da_borda == 'amarelo':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        if cor == 'vermelho' and cor_da_borda == 'preto':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'vermelho' and cor_da_borda == 'branco':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'vermelho' and cor_da_borda == 'vermelho':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'vermelho' and cor_da_borda == 'azul':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'vermelho' and cor_da_borda == 'verde':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'vermelho' and cor_da_borda == 'amarelo':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        if cor == 'azul' and cor_da_borda == 'preto':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'azul' and cor_da_borda == 'branco':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'azul' and cor_da_borda == 'vermelho':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'azul' and cor_da_borda == 'azul':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'azul' and cor_da_borda == 'verde':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'azul' and cor_da_borda == 'amarelo':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        if cor == 'verde' and cor_da_borda == 'preto':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'verde' and cor_da_borda == 'branco':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'verde' and cor_da_borda == 'vermelho':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'verde' and cor_da_borda == 'azul':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'verde' and cor_da_borda == 'verde':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'verde' and cor_da_borda == 'amarelo':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        if cor == 'amarelo' and cor_da_borda == 'preto':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'amarelo' and cor_da_borda == 'branco':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'amarelo' and cor_da_borda == 'vermelho':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'amarelo' and cor_da_borda == 'azul':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'amarelo' and cor_da_borda == 'verde':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'amarelo' and cor_da_borda == 'amarelo':
            plt.hist(lista_organizada, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        plt.title(f'{titulo}')

        plt.xlabel(f'{titulo_eixox}')
        plt.ylabel(f'{titulo_eixoy}')

        plt.show()

    def grafico_de_linha():
        cor = input('Digite a cor do gráfico: ')

        cor_ingles_preto = cor.replace(f'{cor}', 'black')
        cor_ingles_branco = cor.replace(f'{cor}', 'white')
        cor_ingles_vermelho = cor.replace(f'{cor}', 'red')
        cor_ingles_azul = cor.replace(f'{cor}', 'blue')
        cor_ingles_verde = cor.replace(f'{cor}', 'green')
        cor_ingles_amarelo = cor.replace(f'{cor}', 'yellow')

        titulo = input('Digite o título do gráfico: ')

        titulo_eixox = input('Digite o título do eixo x: ')
        titulo_eixoy = input('Digite o título do eixo y: ')

        if cor == 'preto':
            plt.plot(lista, linestyle='--', color=f'{cor_ingles_preto}', label='dados')
        if cor == 'branco':
            plt.plot(lista, linestyle='--', color=f'{cor_ingles_branco}', label='dados')
        if cor == 'vermelho':
            plt.plot(lista, linestyle='--', color=f'{cor_ingles_vermelho}', label='dados')
        if cor == 'azul':
            plt.plot(lista, linestyle='--', color=f'{cor_ingles_azul}', label='dados')
        if cor == 'verde':
            plt.plot(lista, linestyle='--', color=f'{cor_ingles_verde}', label='dados')
        if cor == 'amarelo':
            plt.plot(lista, linestyle='--', color=f'{cor_ingles_amarelo}', label='dados')

        plt.title(f'{titulo}')

        plt.xlabel(f'{titulo_eixox}')
        plt.ylabel(f'{titulo_eixoy}')

        plt.legend()
        plt.grid(True)

        plt.show()

    qtde_de_dados = int(input('Digite a quantidade de dados a ser tratada:'))

    print('use o . ao invés da , para separar as casas decimais dos dados')

    # Lê uma quantidade de dados definida pelo usuário, faz a soma, determina a quantidade de dados digitados e
    # adiciona em uma lista

    for i in range(1, qtde_de_dados+1):
        numero = float(input('Digite um número:'))
        soma += numero
        contador += 1
        lista.append(numero)

    # Calcula a média dos dados repassados

    media = soma/contador

    # Faz a conta da variância sem a raíz quadrada

    for n in range(0, qtde_de_dados):
        variancia_sem_quadrado += ((lista[n]-media)**2)/(qtde_de_dados-1)

    # Tira a raíz da fórmula anterior, completando a conta da variância

    variancia = variancia_sem_quadrado**(1/2)

    # Calcula o erro da média

    erro_da_media = variancia/(qtde_de_dados**(1/2))

    lista_organizada = sorted(lista)

    if contador % 2 == 1:
        m += int((qtde_de_dados+1)/2)
        mediana = lista_organizada[m-1]

    else:
        mediana = (lista[int((qtde_de_dados/2)-1)]+lista[int(qtde_de_dados/2)])/2

    print(f'Esta é a lista com os dados: {lista}')
    print(f'Esta é a quantidade de dados: {contador}')
    print(f'Esta é a soma de todos os dados: {soma}')
    print(f'Esta é a média dos dados: {media}')
    print(f'Esta é a variância: {variancia}')
    print(f'Este é o erro da média: {erro_da_media}')
    print(f'Esta é a mediana da lista: {mediana}')

    # Faz o gráfico usando a função definida anteriormente no início deste bloco

    pergunta_grafico = input('Deseja saber o gráfico desta lista? \n').lower()

    if pergunta_grafico == 'sim' or 'com certeza' or 'claro' or 'óbvio' or 'obvio':
        qual_grafico = input('Digite o nome do gráfico que deseja obter: \n-histograma \n'
                             '-grafico de linhas\n')

        if qual_grafico == 'histograma':
            histograma()
        if qual_grafico == 'grafico de linhas':
            grafico_de_linha()

    elif pergunta_grafico == 'não' or 'nao':
        print('Até mais')

# Puxa os dados de uma base de dados fora do pycharm


def puxar_dados():
    lista_elemento = []
    variancia_sem_quadrado = 0
    contador = 0
    cont_none = 0
    soma = 0
    m = 0

    def histograma():
        cor = input('Digite a cor do gráfico: ')

        cor_ingles_preto = cor.replace(f'{cor}', 'black')
        cor_ingles_branco = cor.replace(f'{cor}', 'white')
        cor_ingles_vermelho = cor.replace(f'{cor}', 'red')
        cor_ingles_azul = cor.replace(f'{cor}', 'blue')
        cor_ingles_verde = cor.replace(f'{cor}', 'green')
        cor_ingles_amarelo = cor.replace(f'{cor}', 'yellow')

        cor_da_borda = input('Digite a cor da borda: ')

        cor_da_borda_ingles_preto = cor.replace(f'{cor}', 'black')
        cor_da_borda_ingles_branco = cor.replace(f'{cor}', 'white')
        cor_da_borda_ingles_vermelho = cor.replace(f'{cor}', 'red')
        cor_da_borda_ingles_azul = cor.replace(f'{cor}', 'blue')
        cor_da_borda_ingles_verde = cor.replace(f'{cor}', 'green')
        cor_da_borda_ingles_amarelo = cor.replace(f'{cor}', 'yellow')

        titulo = input('Digite o título do gráfico: ')
        titulo_eixox = input('Digite o título do eixo x: ')
        titulo_eixoy = input('Digite o título do eixo y: ')

        if cor == 'preto' and cor_da_borda == 'preto':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'preto' and cor_da_borda == 'branco':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'preto' and cor_da_borda == 'vermelho':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'preto' and cor_da_borda == 'azul':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'preto' and cor_da_borda == 'verde':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'preto' and cor_da_borda == 'amarelo':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_preto}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        if cor == 'branco' and cor_da_borda == 'preto':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'branco' and cor_da_borda == 'branco':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'branco' and cor_da_borda == 'vermelho':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'branco' and cor_da_borda == 'azul':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'branco' and cor_da_borda == 'verde':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'branco' and cor_da_borda == 'amarelo':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_branco}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        if cor == 'vermelho' and cor_da_borda == 'preto':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'vermelho' and cor_da_borda == 'branco':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'vermelho' and cor_da_borda == 'vermelho':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'vermelho' and cor_da_borda == 'azul':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'vermelho' and cor_da_borda == 'verde':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'vermelho' and cor_da_borda == 'amarelo':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_vermelho}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        if cor == 'azul' and cor_da_borda == 'preto':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'azul' and cor_da_borda == 'branco':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'azul' and cor_da_borda == 'vermelho':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'azul' and cor_da_borda == 'azul':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'azul' and cor_da_borda == 'verde':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'azul' and cor_da_borda == 'amarelo':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_azul}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        if cor == 'verde' and cor_da_borda == 'preto':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'verde' and cor_da_borda == 'branco':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'verde' and cor_da_borda == 'vermelho':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'verde' and cor_da_borda == 'azul':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'verde' and cor_da_borda == 'verde':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'verde' and cor_da_borda == 'amarelo':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_verde}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        if cor == 'amarelo' and cor_da_borda == 'preto':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_preto}')
        if cor == 'amarelo' and cor_da_borda == 'branco':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_branco}')
        if cor == 'amarelo' and cor_da_borda == 'vermelho':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_vermelho}')
        if cor == 'amarelo' and cor_da_borda == 'azul':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_azul}')
        if cor == 'amarelo' and cor_da_borda == 'verde':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_verde}')
        if cor == 'amarelo' and cor_da_borda == 'amarelo':
            plt.hist(keyword_list, bins=contador, color=f'{cor_ingles_amarelo}',
                     edgecolor=f'{cor_da_borda_ingles_amarelo}')

        plt.title(f'{titulo}')

        plt.xlabel(f'{titulo_eixox}')
        plt.ylabel(f'{titulo_eixoy}')

        plt.show()

    def grafico_de_linha():
        cor = input('Digite a cor do gráfico: ')

        cor_ingles_preto = cor.replace(f'{cor}', 'black')
        cor_ingles_branco = cor.replace(f'{cor}', 'white')
        cor_ingles_vermelho = cor.replace(f'{cor}', 'red')
        cor_ingles_azul = cor.replace(f'{cor}', 'blue')
        cor_ingles_verde = cor.replace(f'{cor}', 'green')
        cor_ingles_amarelo = cor.replace(f'{cor}', 'yellow')

        titulo = input('Digite o título do gráfico: ')

        titulo_eixox = input('Digite o título do eixo x: ')
        titulo_eixoy = input('Digite o título do eixo y: ')

        if cor == 'preto':
            plt.plot(keyword_list, linestyle='--', color=f'{cor_ingles_preto}', label='dados')
        if cor == 'branco':
            plt.plot(keyword_list, linestyle='--', color=f'{cor_ingles_branco}', label='dados')
        if cor == 'vermelho':
            plt.plot(keyword_list, linestyle='--', color=f'{cor_ingles_vermelho}', label='dados')
        if cor == 'azul':
            plt.plot(keyword_list, linestyle='--', color=f'{cor_ingles_azul}', label='dados')
        if cor == 'verde':
            plt.plot(keyword_list, linestyle='--', color=f'{cor_ingles_verde}', label='dados')
        if cor == 'amarelo':
            plt.plot(keyword_list, linestyle='--', color=f'{cor_ingles_amarelo}', label='dados')

        plt.title(f'{titulo}')

        plt.xlabel(f'{titulo_eixox}')
        plt.ylabel(f'{titulo_eixoy}')

        plt.legend()
        plt.grid(True)

        plt.show()

    # lê e ativa o arquivo excel.

    nome_arquivo = input('Digite o nome junto com a extensão dele:\n')
    arquivo = load_workbook(filename=f'{nome_arquivo}')
    arquivo_carregado = arquivo.active

    # armazena em variáveis o número máximo de linhas e colunas do arquivo.

    coluna = arquivo_carregado.max_column
    linha = arquivo_carregado.max_row

    # pega o elemento de cada linha e coluna e armazena em uma lista do python

    for j in range(1, coluna + 1):
        for i in range(1, linha + 1):
            cell_obj = arquivo_carregado.cell(row=i, column=j)
            lista_elemento.append(cell_obj.value)
            contador += 1

            # armazena a quantidade de Nones presentes no arquivo.

            if cell_obj.value is None:
                cont_none += 1

    # elimina os Nones da lista.

    keyword_list = list(filter(None, lista_elemento))

    # elimina a quantidade de Nones da quantidade total de elementos.

    cont_final = contador-cont_none

    # soma todos os elementos da lista.

    for i in range(0, len(keyword_list)):
        soma += keyword_list[i]

    # calcula a média dos elementos da lista.

    media = soma/cont_final

    # faz o cálculo da parte de dentro da raíz quadrada.

    for i in range(0, len(keyword_list)):
        variancia_sem_quadrado += ((keyword_list[i]-media)**2)/(cont_final-1)

    # tira a raíz dá fórmula anterior, chegando na variância. E depois na fórmula do erro da média.

    variancia = variancia_sem_quadrado**(1/2)
    erro_da_media = variancia/(cont_final**(1/2))

    # organiza a lista

    lista_organizada = sorted(keyword_list)

    # pega a mediana da lista

    if contador % 2 == 1:
        m += int((cont_final + 1) / 2)
        mediana = lista_organizada[m - 1]

    else:
        mediana = (lista_organizada[int((cont_final / 2) - 1)] + lista_organizada[int(cont_final / 2)]) / 2

    print(f'A lista com os elementos é: {keyword_list}')
    print(f'O número de elementos é: {cont_final}')
    print(f'A soma é: {soma}')
    print(f'A média é: {media}')
    print(f'A variância é: {variancia}')
    print(f'O erro da média é: {erro_da_media}')
    print(f'A mediana da lista é: {mediana}')

    pergunta_grafico = input('Deseja saber o gráfico desta lista? \n').lower()

    if pergunta_grafico == 'sim' or 'com certeza' or 'claro' or 'óbvio' or 'obvio':
        qual_grafico = input('Digite o nome do gráfico que deseja obter: \n-histograma \n'
                             '-grafico de linhas\n')

        if qual_grafico == 'histograma':
            histograma()
        if qual_grafico == 'grafico de linhas':
            grafico_de_linha()

    elif pergunta_grafico == 'não' or 'nao':
        print('Até mais')


print('-' * 23)
print('auxiador de relatórios')
print('-' * 23)

caminho = input('Como deseja tratar os dados:\n'
                'digite 1 para digitar os dados manualmente\n'
                'digite 2 para puxar os dados de uma arquivo\n')

if caminho == '1':
    digitar_dados()
elif caminho == '2':
    puxar_dados()
